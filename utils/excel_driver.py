import os
import datetime
from openpyxl import Workbook, load_workbook

ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
SHEET_NAME = "Test Data"


class ExcelDriver:
    def __init__(self):
        date_now = datetime.datetime.now()
        date = str(date_now.year) + str(date_now.month) + str(date_now.day)
        self.file_name = f"{ROOT_DIR}\\..\\output\\TestData_{date}.xlsx"

    def set_workbook(self):
        if not os.path.isfile(self.file_name):
            wb = Workbook()
            ws = wb.worksheets[0]
            ws.title = SHEET_NAME
            wb.save(self.file_name)
        self.add_row()

    def add_row(self):
        wb = load_workbook(self.file_name)
        ws = wb[SHEET_NAME]
        if ws.cell(row=ws.max_row, column=1).value != "END":
            ws.cell(row=ws.max_row + 1, column=1, value="END")
        wb.save(self.file_name)

    def set_value(self, col_name, cell_value):
        col_num = self.get_column(col_name)
        wb = load_workbook(self.file_name)
        ws = wb[SHEET_NAME]
        ws.cell(row=ws.max_row, column=col_num).value = cell_value
        wb.save(self.file_name)

    def get_column(self, col_name):
        wb = load_workbook(self.file_name)
        ws = wb[SHEET_NAME]
        col_num = 0
        for cell in ws[1]:
            if cell.value == col_name:
                col_num = cell.column
                break
        if col_num == 0:
            ws.cell(row=1, column=ws.max_column + 1, value=None)
            for cell in ws[1]:
                if cell.value is None:
                    cell.value, col_num = col_name, cell.column
                    break
        wb.save(self.file_name)
        return col_num
